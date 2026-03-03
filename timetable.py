#!/usr/bin/env python3
"""Generate an Excel timetable with formatting and overlap detection."""
from datetime import datetime
import re

schedule = [
    ("09:00~09:30", "報到", ""),
    ("09:30~09:40", "開場致詞", "王教授"),
    ("09:40~10:05", "從 MWC 2026 到 6G：AI-RAN 標準、開源與互通測試的最新進展", "劉教授"),
    ("10:05~10:30", "基於 O-RAN 開放介面的 ISAC 無線感知技術", "陳教授"),
    ("10:30~10:50", "Break", ""),
    ("10:50~11:20", "Federated Foundational Models in AI-RAN：Practical and Forward Looking Perspective", "教學團隊"),
    ("11:20~12:00", "O-RAN 環境與各模組化功能介紹", "教學團隊"),
    ("12:00~13:30", "Lunch", ""),
    ("13:30~14:00", "O-RAN 開源軟體組織簡介", "教學團隊"),
    ("14:00~14:30", "O-RAN 實驗環境建置教學", "教學團隊"),
    ("14:30~14:50", "Break", ""),
    ("14:50~15:50", "O-RAN xApps 實作建置教學", "教學團隊"),
    ("15:50~16:30", "現場討論時間", "教學團隊"),
]

time_format = "%H:%M"

def parse_time_range(rng):
    start_str, end_str = rng.split("~")
    return datetime.strptime(start_str, time_format), datetime.strptime(end_str, time_format)


def detect_conflicts(entries):
    # entries: list of (start,end,description)
    conflicts = []
    sorted_entries = sorted(entries, key=lambda x: x[0])
    for i in range(len(sorted_entries) - 1):
        cur_start, cur_end, cur_desc = sorted_entries[i]
        nxt_start, nxt_end, nxt_desc = sorted_entries[i + 1]
        if cur_end > nxt_start:
            conflicts.append((sorted_entries[i], sorted_entries[i + 1]))
    return conflicts


def create_excel(filepath='timetable.xlsx'):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Side, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = 'Schedule'

    # build time slots (30-minute intervals) for the day covering earliest to latest
    # collect all unique start/end times from schedule
    times = []
    for tim,_,_ in schedule:
        s,e = parse_time_range(tim)
        times.append(s)
        times.append(e)
    if not times:
        wb.save(filepath)
        return
    times = sorted(set(times))

    # header row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.cell(row=1, column=1, value="Time").alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=1, column=3, value="Content").alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # build index map using strings to avoid datetime mismatches
    times_str = [t.strftime(time_format) for t in times]
    index_map = {times_str[i]: i for i in range(len(times_str))}

    # fill rows with consecutive intervals
    row = 2
    slot_map = {}  # map start time string to row number
    for i in range(len(times) - 1):
        start = times_str[i]
        end = times_str[i+1]
        ws.cell(row=row, column=1, value=start)
        ws.cell(row=row, column=2, value=end)
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).border = border
        slot_map[start] = row
        row += 1

    # place entries with merges and formatting
    conflicts = []
    for tim, content, speaker in schedule:
        start, end = parse_time_range(tim)
        s_str = start.strftime(time_format)
        e_str = end.strftime(time_format)
        start_row = slot_map[s_str]
        start_idx = index_map[s_str]
        end_idx = index_map[e_str]
        span = end_idx - start_idx
        end_row = start_row + span - 1
        cell = ws.cell(row=start_row, column=3, value=f"{content}\n{speaker}")
        if span > 1:
            ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # conflict detection
    parsed = []
    for tim, content, speaker in schedule:
        st, en = parse_time_range(tim)
        parsed.append((st, en, f"{content} ({speaker})"))
    conflicted = detect_conflicts(parsed)
    if conflicted:
        print("Conflicts detected:")
        fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        for (s1,e1,d1),(s2,e2,d2) in conflicted:
            print(f"  {s1.strftime(time_format)}~{e1.strftime(time_format)} '{d1}' overlaps {s2.strftime(time_format)}~{e2.strftime(time_format)} '{d2}'")
            # color all cells in the merged range for both entries
            for start_time, end_time in ((s1,e1),(s2,e2)):
                sstr = start_time.strftime(time_format)
                estr = end_time.strftime(time_format)
                rstart = slot_map[sstr]
                idx1 = index_map[sstr]
                idx2 = index_map[estr]
                span_cells = idx2 - idx1
                for r in range(rstart, rstart + span_cells):
                    ws.cell(row=r, column=3).fill = fill

    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    wb.save(filepath)
    print(f"Excel written to {filepath}")


if __name__ == '__main__':
    create_excel()
