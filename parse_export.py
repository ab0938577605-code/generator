import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from datetime import datetime

# 原始資料輸入
data_text = """
時間：09:00~09:30，內容：報到，講者：無
時間：09:30~09:40，內容：開場致詞，講者：王教授
時間：09:40~10:05，內容：從 MWC 2026 到 6G：AI-RAN 標準、開源與互通測試的最新進展，講者：劉教授
時間：10:05~10:30，內容：基於 O-RAN 開放介面的 ISAC 無線感知技術，講者：陳教授
時間：10:30~10:50，內容：Break，講者：無
時間：10:50~11:20，內容：Federated Foundational Models in AI-RAN：Practical and Forward Looking Perspective，講者：教學團隊
時間：11:20~12:00，內容：O-RAN 環境與各模組化功能介紹，講者：教學團隊
時間：12:00~13:30，內容：Lunch，講者：無
時間：13:30~14:00，內容：O-RAN 開源軟體組織簡介，講者：教學團隊
時間：14:00~14:30，內容：O-RAN 實驗環境建置教學，講者：教學團隊
時間：14:30~14:50，內容：Break，講者：無
時間：14:50~15:50，內容：O-RAN xApps 實作建置教學，講者：教學團隊
時間：15:50~16:30，內容：現場討論時間，講者：教學團隊
"""

def parse_and_export(text):
    # 1. 解析資料
    pattern = r"時間：(.*?)~(.*?)，內容：(.*?)，講者：(.*?)$"
    rows = []
    lines = [l.strip() for l in text.strip().split('\n') if l.strip()]
    
    for line in lines:
        match = re.search(pattern, line)
        if match:
            rows.append({
                "start": match.group(1),
                "end": match.group(2),
                "content": match.group(3),
                "speaker": match.group(4)
            })

    # 2. 衝突檢查
    conflicts = []
    for i in range(len(rows)):
        for j in range(i + 1, len(rows)):
            s1, e1 = rows[i]['start'], rows[i]['end']
            s2, e2 = rows[j]['start'], rows[j]['end']
            if max(s1, s2) < min(e1, e2): # 時間重疊邏輯
                print(f"⚠️ 衝突警告: [{s1}-{e1} {rows[i]['content']}] 與 [{s2}-{e2} {rows[j]['content']}] 重疊！")
                conflicts.extend([i, j])

    # 3. 建立 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"
    
    # 設定表頭
    ws.append(["Time", "Content", "Speaker"])
    
    # 樣式準備
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # 寫入內容
    for idx, row in enumerate(rows):
        display_time = f"{row['start']}-{row['end']}"
        # 處理換行格式：內容 \n 講者 (若講者為無則只顯示內容)
        display_content = row['content']
        display_speaker = row['speaker']
        
        excel_row = idx + 2
        ws.cell(row=excel_row, column=1, value=display_time)
        ws.cell(row=excel_row, column=2, value=display_content)
        ws.cell(row=excel_row, column=3, value=display_speaker)

        # 衝突變紅
        if idx in conflicts:
            for col in range(1, 4):
                ws.cell(row=excel_row, column=col).fill = red_fill

    # 4. 合併儲存格邏輯 (檢查連續列是否為同一講者/課程)
    # 這裡實作範例圖中「教學團隊」跨多時段合併
    curr_row = 2
    while curr_row <= len(rows) + 1:
        next_row = curr_row + 1
        # 如果下一列的講者跟內容與這一列相同，則考慮合併
        while next_row <= len(rows) + 1 and \
              ws.cell(row=next_row, column=3).value == ws.cell(row=curr_row, column=3).value and \
              ws.cell(row=curr_row, column=3).value != "無":
            next_row += 1
        
        if next_row > curr_row + 1:
            ws.merge_cells(start_row=curr_row, start_column=3, end_row=next_row-1, end_column=3)
        curr_row = next_row

    # 5. 最後格式調整
    for row in ws.iter_rows(min_row=1, max_row=len(rows)+1, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = alignment
            cell.border = thin_border

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15

    wb.save("timetable.xlsx")
    print("\n✅ Excel 已成功輸出至當前資料夾：timetable.xlsx")

if __name__ == "__main__":
    parse_and_export(data_text)